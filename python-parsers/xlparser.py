#!/usr/bin/env python3

#
# This uses the openpyxl module, which can be installed using
#
# pip install openpyxl
#
# To run:
#
# xlparser.py filename >some_file.csv
#

from openpyxl import load_workbook
import sys
import re

def read_document_map(workbook):
    ws_document_map = workbook['Document map']

    cells = ws_document_map["A:C"]
    row_count = len(cells[2])

    dmap = {}

    precinct_code = None
    for i in range(1,row_count):
        b = cells[1][i]
        if b.value is not None:
            precinct_code = b.value

        c = cells[2][i]
        if c.value is not None:
            if precinct_code not in dmap:
                dmap[precinct_code] = { 'precinct': precinct_code, 'offices': [] }

            dmap[precinct_code]['offices'].append(c.value)

    return dmap

def read_precinct_results(worksheet, precinct_offices, county):
    cells = worksheet["A:AB"]
    precinct_code = cells[0][24].value

    results = []
    for i in range(28, len(cells[0])):
        v = cells[0][i]
        print(v.value)
        if v.value in precinct_offices:
            office = v.value
            if " - Democratic Party" in office:
                office_party = "DEM"
                office = office.replace(" - Democratic Party", "")
            elif " - Republican Party" in office:
                office_party = "REP"
                office = office.replace(" - Republican Party", "")

            dist = ''
            if "District " in office:
                # strip district number from office and move it to "dist" field
                m = re.search(", District No. ([0-9]*)", office)
                pattern = "District No."
                if m is None:
                    m = re.search(", District ([0-9]*)", office)
                    pattern = "District"
                if m is not None:
                    dist = m.group(1)
                    str = ", %s %s" % (pattern, dist)
                    office = office.replace(str, '')

            i = i + 2
            v = cells[0][i]
            while not "Choice" in v.value:
                print(cells[26][i].value)
                choice = v.value
                party = cells[5][i].value
                if party == "":
                    party = office_party
                absentee = cells[16][i].value
                early = cells[10][i].value
                election_day = cells[19][i].value
                total = cells[26][i].value

                print("%s,%s,\"%s\",%s,%s,%s,%d,%d,%d,%d,%d" % (county, precinct_code, office, dist, choice, party, total, early, election_day, absentee, 0))

                i = i + 1
                v = cells[0][i]

filename = sys.argv[1]

workbook = load_workbook(filename=filename)

document_map = read_document_map(workbook)

print("county,precinct,office,district,candidate,party,votes,early_voting,election_day,mail,provisional")
for name in workbook.sheetnames[1:-1]:
    worksheet = workbook[name]
    title = worksheet["A4"].value
    if "Precinct Results" in title:
        location = worksheet["N2"].value
        if "COUNTY" in location:
            m = re.search("([A-Za-z]*) COUNTY", location)
            location = m.group(1)

        precinct_code = worksheet["A28"].value
        read_precinct_results(worksheet, document_map[precinct_code]['offices'], location)
